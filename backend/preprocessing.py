import io
import warnings
import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")


# KONSTANTA
SHEET_NAME = "BFTP"

FEATURE_COLS = {
    # === IDENTITAS ===
    "Material description":  "material_desc",
    "Material":              "material_kode",
    "Code Material ":        "code_material",
    "Batch Ke-":             "batch_ke",

    # === LAPIS 1 - GRANULASI BASAH ===
    "LAPIS 1":                               "lapis1_warna",
    "BULAN PRODUKSI CB 1":                   "cb1_bulan",
    "SHIFT":                                 "cb1_shift",
    "LINE":                                  "cb1_line",
    "JML DECOCT           ( Kg )":           "cb1_jml_decoct_kg",
    "SUHU DECOCT LOT 1/1":                   "cb1_suhu_lot1_1",
    "SUHU DECOCT LOT 1/2":                   "cb1_suhu_lot1_2",
    "SUHU DECOCT LOT 2/1":                   "cb1_suhu_lot2_1",
    "SUHU DECOCT LOT 2/2":                   "cb1_suhu_lot2_2",
    "SUHU DECOCT LOT 3/1":                   "cb1_suhu_lot3_1",
    "SUHU DECOCT LOT 3/2":                   "cb1_suhu_lot3_2",
    "SUHU DECOCT LOT 4/1":                   "cb1_suhu_lot4_1",
    "SUHU DECOCT LOT 4/2":                   "cb1_suhu_lot4_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 1/1":  "cb1_suhu_diosna_lot1_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 1/2":  "cb1_suhu_diosna_lot1_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 2/1":  "cb1_suhu_diosna_lot2_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 2/2":  "cb1_suhu_diosna_lot2_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 3/1":  "cb1_suhu_diosna_lot3_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 3/2":  "cb1_suhu_diosna_lot3_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 4/1":  "cb1_suhu_diosna_lot4_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 4/2":  "cb1_suhu_diosna_lot4_2",
    "JML GRANUL BR (-)  (KG)":               "cb1_jml_granul_kg",
    "TEORITIS GRANUL":                       "cb1_teoritis_granul",
    "% Yield ":                              "cb1_pct_yield",
    "KA 1":                                  "cb1_ka1",
    "KA 2":                                  "cb1_ka2",
    "KA 3":                                  "cb1_ka3",
    "KA 4":                                  "cb1_ka4",
    "RATA    KA":                            "cb1_rata_ka",
    "MIXER 1":                               "cb1_mixer1",
    "MIXER 2":                               "cb1_mixer2",
    "MIXER 3":                               "cb1_mixer3",
    "MIXER 4":                               "cb1_mixer4",
    "CHOPPER 1":                             "cb1_chopper1",
    "CHOPPER 2":                             "cb1_chopper2",
    "CHOPPER 3":                             "cb1_chopper3",
    "CHOPPER 4":                             "cb1_chopper4",
    "Inlet lot 1 Phase 1":                   "cb1_inlet_lot1_ph1",
    "Inlet lot 1 Phase 2":                   "cb1_inlet_lot1_ph2",
    "Inlet lot 2 Phase 1":                   "cb1_inlet_lot2_ph1",
    "Inlet lot 2 Phase 2":                   "cb1_inlet_lot2_ph2",
    "Outlet lot 1 Phase 1":                  "cb1_outlet_lot1_ph1",
    "Outlet lot 1 Phase 2":                  "cb1_outlet_lot1_ph2",
    "Outlet lot 2 Phase 1":                  "cb1_outlet_lot2_ph1",
    "Outlet lot 2 Phase 2":                  "cb1_outlet_lot2_ph2",
    "Waktu lot 1 Phase 1":                   "cb1_waktu_lot1_ph1",
    "Waktu lot 1 Phase 2":                   "cb1_waktu_lot1_ph2",
    "Waktu lot 2 Phase 1":                   "cb1_waktu_lot2_ph1",
    "Waktu lot 2 Phase 2":                   "cb1_waktu_lot2_ph2",
    "NO AYAKAN ALEX":                        "cb1_no_ayakan_alex",
    "NO AYAKAN OG YC/ FREWITT":              "cb1_no_ayakan_frewitt",

    # === LAPIS 1 - CAMPUR KERING ===
    "JML GRANUL (+) NYATA ( Kg )":           "ck1_jml_granul_nyata",
    "TEORITIS GRANUL_1":                     "ck1_teoritis_granul",
    "% YIELD":                               "ck1_pct_yield",
    "KA setelah CK":                         "ck1_ka_setelah_ck",
    "KETERANGAN":                            "ck1_keterangan",

    # === LAPIS 2 - GRANULASI BASAH ===
    "LAPIS 2":                               "lapis2_warna",
    "BULAN PRODUKSI CB 2":                   "cb2_bulan",
    "SHIFT_1":                               "cb2_shift",
    "LINE_2":                                "cb2_line",
    "JML DECOCT           ( Kg )_1":         "cb2_jml_decoct_kg",
    "SUHU DECOCT LOT 1/1_1":                "cb2_suhu_lot1_1",
    "SUHU DECOCT LOT 1/2_1":                "cb2_suhu_lot1_2",
    "SUHU DECOCT LOT 2/1_1":                "cb2_suhu_lot2_1",
    "SUHU DECOCT LOT 2/2_1":                "cb2_suhu_lot2_2",
    "SUHU DECOCT LOT 3/1_1":                "cb2_suhu_lot3_1",
    "SUHU DECOCT LOT 3/2_1":                "cb2_suhu_lot3_2",
    "SUHU DECOCT LOT 4/1_1":                "cb2_suhu_lot4_1",
    "SUHU DECOCT LOT 4/2_1":                "cb2_suhu_lot4_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 1/1_1": "cb2_suhu_diosna_lot1_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 1/2_1": "cb2_suhu_diosna_lot1_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 2/1_1": "cb2_suhu_diosna_lot2_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 2/2_1": "cb2_suhu_diosna_lot2_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 3/1_1": "cb2_suhu_diosna_lot3_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 3/2_1": "cb2_suhu_diosna_lot3_2",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 4/1_1": "cb2_suhu_diosna_lot4_1",
    "SUHU DECOCT SEBELUM MASUK DIOSNA LOT 4/2_1": "cb2_suhu_diosna_lot4_2",
    "JML GRANUL BR (-)  (KG)_1":             "cb2_jml_granul_kg",
    "TEORITIS GRANUL_2":                     "cb2_teoritis_granul",
    "% YILD":                                "cb2_pct_yield",
    "KA 1_1":                                "cb2_ka1",
    "KA 2_1":                                "cb2_ka2",
    "KA 3_1":                                "cb2_ka3",
    "KA 4_1":                                "cb2_ka4",
    "RATA-RATA KA":                          "cb2_rata_ka",
    "MIXER 1_1":                             "cb2_mixer1",
    "MIXER 2_1":                             "cb2_mixer2",
    "MIXER 3_1":                             "cb2_mixer3",
    "MIXER 4_1":                             "cb2_mixer4",
    "CHOPPER 1_1":                           "cb2_chopper1",
    "CHOPPER 2_1":                           "cb2_chopper2",
    "CHOPPER 3_1":                           "cb2_chopper3",
    "CHOPPER 4_1":                           "cb2_chopper4",
    "Inlet lot 1 Phase 1_1":                 "cb2_inlet_lot1_ph1",
    "Inlet lot 1 Phase 2_1":                 "cb2_inlet_lot1_ph2",
    "Inlet lot 2 Phase 1_1":                 "cb2_inlet_lot2_ph1",
    "Inlet lot 2 Phase 2_1":                 "cb2_inlet_lot2_ph2",
    "Inlet lot 3 Phase 1_1":                 "cb2_inlet_lot3_ph1",
    "Inlet lot 3 Phase 2_1":                 "cb2_inlet_lot3_ph2",
    "Inlet lot 4 Phase 1_1":                 "cb2_inlet_lot4_ph1",
    "Inlet lot 4  Phase 2":                  "cb2_inlet_lot4_ph2",
    "Outlet lot 1 Phase 1_1":               "cb2_outlet_lot1_ph1",
    "Outlet lot 1 Phase 2_1":               "cb2_outlet_lot1_ph2",
    "Outlet lot 2 Phase 1_1":               "cb2_outlet_lot2_ph1",
    "Outlet lot 2  phase 2":                 "cb2_outlet_lot2_ph2",
    "outlet lot 3 Phase 1_1":               "cb2_outlet_lot3_ph1",
    "outlet lot 3 Phase 2_1":               "cb2_outlet_lot3_ph2",
    "Outlet lot 4 Phase 1_1":               "cb2_outlet_lot4_ph1",
    "Outlet lot 4 Phase 2_1":               "cb2_outlet_lot4_ph2",
    "Waktu lot 1 Phase 1_1":                 "cb2_waktu_lot1_ph1",
    "Waktu lot 1 Phase 2_1":                 "cb2_waktu_lot1_ph2",
    "Waktu lot 2 Phase 1_1":                 "cb2_waktu_lot2_ph1",
    "Waktu lot 2 Phase 2_1":                 "cb2_waktu_lot2_ph2",
    "Waktu lot 3 Phase 1_1":                 "cb2_waktu_lot3_ph1",
    "Waktu lot 3 Phase 2_1":                 "cb2_waktu_lot3_ph2",
    "waktu lot 4 Phase 1_1":                 "cb2_waktu_lot4_ph1",
    "waktu lot 4 Phase 2_1":                 "cb2_waktu_lot4_ph2",
    "NO AYAKAN ALEX_1":                      "cb2_no_ayakan_alex",
    "NO AYAKAN OG YC/ FREWITT_1":            "cb2_no_ayakan_frewitt",

    # === LAPIS 2 - CAMPUR KERING ===
    "JML GRANUL BR (+) - (KG)":              "ck2_jml_granul_nyata",
    "TEORITIS GRANULASI KEING":              "ck2_teoritis_granul",
    "% YILD_1":                              "ck2_pct_yield",
    "KA SETELAH CK":                         "ck2_ka_setelah_ck",
    "KETERANGAN_1":                          "ck2_keterangan",
}

CATEGORICAL_COLS = [
    "material_desc", "material_kode", "code_material",
    "lapis1_warna", "lapis2_warna",
    "cb1_bulan", "cb2_bulan",
    "cb1_line", "cb2_line",
    "ck1_keterangan", "ck2_keterangan",
    "cb1_no_ayakan_alex", "cb1_no_ayakan_frewitt",
    "cb2_no_ayakan_alex", "cb2_no_ayakan_frewitt",
    "is_defect", "excel_row",
]

KA_COLS = [
    "cb1_ka1", "cb1_ka2", "cb1_ka3", "cb1_ka4",
    "cb2_ka1", "cb2_ka2", "cb2_ka3", "cb2_ka4",
    "ck1_ka_setelah_ck", "ck2_ka_setelah_ck",
]
KA_MIN, KA_MAX = 2.0, 10.0

SUHU_MIN, SUHU_MAX = 55.0, 95.0        # batas wajar suhu
SUHU_THRESHOLD_TINGGI = 75.0            # flag suhu tinggi

BULAN_MAP = {
    "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
    "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
    "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12,
}


# HELPER INTERNAL
def _to_bytes(uploaded_file):
    """Konversi uploaded_file (Streamlit UploadedFile atau path) ke bytes."""
    if isinstance(uploaded_file, (str, bytes)):
        if isinstance(uploaded_file, str):
            with open(uploaded_file, "rb") as f:
                return f.read()
        return uploaded_file
    # Streamlit UploadedFile
    uploaded_file.seek(0)
    return uploaded_file.read()


def _detect_yellow_rows(file_bytes: bytes) -> set:
    """
    Langkah 1a: Deteksi baris Excel yang diwarnai kuning (FFFFFF00 atau FFFF0000).
    Mengembalikan set nomor baris Excel (1-based).
    """
    wb_fmt = load_workbook(io.BytesIO(file_bytes))
    ws_fmt = wb_fmt[SHEET_NAME]
    yellow_rows = set()
    for i, row in enumerate(ws_fmt.iter_rows(min_row=4), start=4):
        cell_b = row[1]  # kolom B = Material description
        fill = cell_b.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb.upper()
            if rgb in ("FFFFFF00", "FFFF0000"):
                yellow_rows.add(i)
    return yellow_rows


def _load_data(file_bytes: bytes, yellow_rows: set):
    """
    Langkah 1b–3: Load workbook data_only, baca header baris 3, baca data ab baris 4.
    Mengembalikan df_raw dengan kolom is_defect & excel_row.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[SHEET_NAME]

    # Baca header baris ke-3
    raw_headers = [cell.value for cell in list(ws.iter_rows(min_row=3, max_row=3, max_col=193))[0]]

    # Rename kolom duplikat dengan suffix
    seen = {}
    clean_headers = []
    for h in raw_headers:
        if h is None:
            h = "UNNAMED"
        if h in seen:
            seen[h] += 1
            clean_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            clean_headers.append(h)

    # Baca baris data (ab baris 4)
    rows_data, is_defect_flags, excel_row_nums = [], [], []
    for i, row in enumerate(ws.iter_rows(min_row=4, max_col=193, values_only=True), start=4):
        vals = list(row)
        if all(v is None for v in vals):
            continue
        # Ganti string formula dengan None
        processed = [None if (isinstance(v, str) and v.startswith("=")) else v for v in vals]
        rows_data.append(processed)
        is_defect_flags.append(1 if i in yellow_rows else 0)
        excel_row_nums.append(i)

    df_raw = pd.DataFrame(rows_data, columns=clean_headers)
    df_raw["is_defect"] = is_defect_flags
    df_raw["excel_row"] = excel_row_nums
    return df_raw


def _select_columns(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Langkah 3: Seleksi & rename kolom sesuai FEATURE_COLS."""
    available = {k: v for k, v in FEATURE_COLS.items() if k in df_raw.columns}
    df = df_raw[list(available.keys()) + ["is_defect", "excel_row"]].copy()
    df = df.rename(columns=available)
    return df


def _clean_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Langkah 4: Drop baris tanpa material_desc & duplikat penuh."""
    df = df.dropna(subset=["material_desc"])
    df = df.drop_duplicates(subset=[c for c in df.columns if c != "excel_row"])
    return df.reset_index(drop=True)


def _handle_ka_outliers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Langkah 5A: Buat flag KA ekstrem SEBELUM normalisasi.
    Nilai KA > 10 atau < 2 dijadikan NaN SETELAH flag disimpan.
    """
    for col in KA_COLS:
        if col not in df.columns:
            continue
        flag_col = col + "_ekstrem"
        df[flag_col] = ((df[col] < KA_MIN) | (df[col] > KA_MAX)).astype(int)
    # Agregat flag
    ka_flag_cols = [c + "_ekstrem" for c in KA_COLS if c in df.columns]
    df["ka_ekstrem_count"] = df[ka_flag_cols].sum(axis=1)
    return df


def _handle_suhu_outliers(df: pd.DataFrame, numeric_cols: list) -> pd.DataFrame:
    """
    Langkah 5B: Set nilai Suhu di luar [55, 95] ke NaN.
    KA outlier TIDAK di-NaN-kan di sini (sudah diflag di 5A).
    """
    suhu_cols = [c for c in numeric_cols if "suhu" in c.lower()]
    for col in suhu_cols:
        if col in df.columns:
            mask = (df[col] < SUHU_MIN) | (df[col] > SUHU_MAX)
            df.loc[mask, col] = np.nan
    return df


def _impute_numeric(df: pd.DataFrame, numeric_cols: list) -> pd.DataFrame:
    """Langkah 5D numerik: median per material → median global."""
    numeric_to_impute = [c for c in numeric_cols if c in df.columns]
    for col in numeric_to_impute:
        if df[col].isnull().sum() == 0:
            continue
        if "material_desc" in df.columns:
            group_median = df.groupby("material_desc")[col].transform("median")
            df[col] = df[col].fillna(group_median)
        global_median = df[col].median()
        df[col] = df[col].fillna(global_median)
    return df


def _impute_categorical(df: pd.DataFrame) -> pd.DataFrame:
    """Langkah 5D kategorik: forward fill + TIDAK DIKETAHUI."""
    for col in ["material_desc", "material_kode", "code_material"]:
        if col in df.columns:
            df[col] = df[col].ffill()
    for col in ["ck1_keterangan", "ck2_keterangan"]:
        if col in df.columns:
            df[col] = df[col].fillna("TIDAK DIKETAHUI")
    for col in ["lapis1_warna", "lapis2_warna"]:
        if col in df.columns:
            df[col] = df[col].ffill()
    return df


def _feature_engineering(df: pd.DataFrame) -> pd.DataFrame:
    """
    Langkah 6: Tambah fitur turunan — PERSIS dari notebook.
    - cb1_suhu_rata, cb2_suhu_rata
    - cb1_suhu_std,  cb2_suhu_std
    - cb1_yield_deviate, ck1_yield_deviate, cb2_yield_deviate
    - cb1_ka_outlier, cb2_ka_outlier
    - cb1_suhu_tinggi, cb2_suhu_tinggi
    - no_batch_num
    """
    # Rata-rata & std suhu decoct L1
    suhu_l1 = [c for c in df.columns if c.startswith("cb1_suhu_lot") and "diosna" not in c]
    if suhu_l1:
        df["cb1_suhu_rata"] = df[suhu_l1].mean(axis=1, skipna=True).round(2)
        df["cb1_suhu_std"]  = df[suhu_l1].std(axis=1,  skipna=True).round(4)

    # Rata-rata & std suhu decoct L2
    suhu_l2 = [c for c in df.columns if c.startswith("cb2_suhu_lot") and "diosna" not in c]
    if suhu_l2:
        df["cb2_suhu_rata"] = df[suhu_l2].mean(axis=1, skipna=True).round(2)
        df["cb2_suhu_std"]  = df[suhu_l2].std(axis=1,  skipna=True).round(4)

    # Yield deviate flags
    if "cb1_pct_yield" in df.columns:
        df["cb1_yield_deviate"] = ((df["cb1_pct_yield"] < 99.0) | (df["cb1_pct_yield"] > 103.1)).astype(int)
    if "ck1_pct_yield" in df.columns:
        df["ck1_yield_deviate"] = ((df["ck1_pct_yield"] < 99.1) | (df["ck1_pct_yield"] > 102.9)).astype(int)
    if "cb2_pct_yield" in df.columns:
        df["cb2_yield_deviate"] = ((df["cb2_pct_yield"] < 99.1) | (df["cb2_pct_yield"] > 102.3)).astype(int)

    # KA outlier flag
    for ka_col in ["cb1_rata_ka", "cb2_rata_ka"]:
        if ka_col in df.columns:
            col_name = ka_col.replace("rata_ka", "ka_outlier")
            df[col_name] = ((df[ka_col] < 4.0) | (df[ka_col] > 6.5)).astype(int)

    # Suhu tinggi flag
    if "cb1_suhu_rata" in df.columns:
        df["cb1_suhu_tinggi"] = (df["cb1_suhu_rata"] > SUHU_THRESHOLD_TINGGI).astype(int)
    if "cb2_suhu_rata" in df.columns:
        df["cb2_suhu_tinggi"] = (df["cb2_suhu_rata"] > SUHU_THRESHOLD_TINGGI).astype(int)

    # Tidak ada batch ke
    if "batch_ke" in df.columns:
        df["no_batch_num"] = df["batch_ke"].isnull().astype(int)

    return df


def _encode_categorical(df: pd.DataFrame) -> pd.DataFrame:
    """Langkah 7: Encode kategorik — PERSIS dari notebook."""
    # Lapis warna
    for col in ["lapis1_warna", "lapis2_warna"]:
        if col in df.columns:
            df[col + "_enc"] = df[col].astype("category").cat.codes

    # Bulan produksi → ordinal
    for col in ["cb1_bulan", "cb2_bulan"]:
        if col in df.columns:
            df[col + "_num"] = df[col].str.upper().map(BULAN_MAP)
            df[col + "_num"] = pd.to_numeric(df[col + "_num"], errors="coerce").fillna(0).astype(int)

    # LINE encode
    for col in ["cb1_line", "cb2_line"]:
        if col in df.columns:
            df[col + "_enc"] = df[col].astype("category").cat.codes

    # Shift → numerik
    for col in ["cb1_shift", "cb2_shift"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    # Keterangan selesai binary
    for col in ["ck1_keterangan", "ck2_keterangan"]:
        if col in df.columns:
            df[col + "_selesai"] = (df[col].str.upper().str.contains("SELESAI", na=False)).astype(int)

    return df


def _rulebased_defect_label(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fallback labeling berbasis threshold — dipakai HANYA ketika tidak ada
    baris kuning di Excel (operator belum menandai secara manual).

    Hanya menggunakan kondisi yang PASTI merupakan indikator defect
    tanpa bergantung pada karakteristik produk:
      - Ada suhu decoct individual > 75°C  (threshold dari 01_preprocessing.ipynb)
      - Yield CB1 < 99.0% atau > 103.1%    (batas spesifikasi)
      - Yield CK1 < 99.1% atau > 102.9%    (batas spesifikasi)
      - Yield CB2 < 99.1% atau > 102.3%    (batas spesifikasi)
      - Tidak ada nomor batch               (baris rework gagal)

    KA sengaja TIDAK dimasukkan karena range normal KA berbeda-beda
    tergantung jenis produk — tidak bisa pakai threshold universal.
    """
    conds = []

    # ── Suhu individual CB1 (threshold 75°C persis dari notebook) ──
    suhu_cb1_cols = [c for c in df.columns
                     if c.startswith("cb1_suhu_lot") and "diosna" not in c]
    for col in suhu_cb1_cols:
        conds.append(df[col] > 75.0)

    # ── Suhu individual CB2 ─────────────────────────────────────────
    suhu_cb2_cols = [c for c in df.columns
                     if c.startswith("cb2_suhu_lot") and "diosna" not in c]
    for col in suhu_cb2_cols:
        conds.append(df[col] > 75.0)

    # ── Yield CB1 (skip 0 = missing) ───────────────────────────────
    if "cb1_pct_yield" in df.columns:
        valid = df["cb1_pct_yield"] > 0
        conds.append(valid & ((df["cb1_pct_yield"] < 99.0) | (df["cb1_pct_yield"] > 103.1)))

    # ── Yield CK1 ──────────────────────────────────────────────────
    if "ck1_pct_yield" in df.columns:
        valid = df["ck1_pct_yield"] > 0
        conds.append(valid & ((df["ck1_pct_yield"] < 99.1) | (df["ck1_pct_yield"] > 102.9)))

    # ── Yield CB2 ──────────────────────────────────────────────────
    if "cb2_pct_yield" in df.columns:
        valid = df["cb2_pct_yield"] > 0
        conds.append(valid & ((df["cb2_pct_yield"] < 99.1) | (df["cb2_pct_yield"] > 102.3)))

    # ── Tidak ada nomor batch (baris rework/sublapis gagal) ─────────
    if "no_batch_num" in df.columns:
        conds.append(df["no_batch_num"] == 1)

    if not conds:
        return df

    # OR semua kondisi → jika salah satu terpenuhi = defect
    combined = conds[0].fillna(False)
    for c in conds[1:]:
        combined = combined | c.fillna(False)

    df["is_defect"] = combined.astype(int)
    return df


def _final_missing_cleanup(df: pd.DataFrame) -> pd.DataFrame:
    """Langkah 8: Isi sisa missing dengan median (numerik) atau UNKNOWN (teks)."""
    mv_final = df.isnull().sum()
    mv_final = mv_final[mv_final > 0]
    for col in mv_final.index:
        if df[col].dtype in [np.float64, np.int64, np.float32, np.int32]:
            df[col] = df[col].fillna(df[col].median())
        else:
            df[col] = df[col].fillna("UNKNOWN")
    return df


# PUBLIC API
def run_preprocessing(uploaded_file, save_path: str = None) -> pd.DataFrame:
    """
    Jalankan seluruh pipeline preprocessing dari 01_preprocessing.ipynb.

    Parameters
    ----------
    uploaded_file : str | bytes | Streamlit UploadedFile
        File Excel mentah (.xlsx / .xls).
    save_path : str, optional
        Jika diberikan, simpan hasil ke path ini sebagai CSV.

    Returns
    -------
    pd.DataFrame
        Dataset bersih siap modeling (= dataset_clean.csv).
    """
    # Konversi ke bytes
    file_bytes = _to_bytes(uploaded_file)

    # 1. Deteksi baris kuning
    yellow_rows = _detect_yellow_rows(file_bytes)

    # 2–3. Load workbook & baca header + data
    df_raw = _load_data(file_bytes, yellow_rows)

    # 3. Seleksi kolom
    df = _select_columns(df_raw)

    # 4. Hapus baris tidak valid
    df = _clean_rows(df)

    # Tentukan kolom numerik
    numeric_cols = [c for c in df.columns if c not in CATEGORICAL_COLS]

    # Konversi ke numerik
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # 5A. Handle KA outlier (buat flag, jangan hapus)
    df = _handle_ka_outliers(df)

    # 5B. Set Suhu outlier → NaN
    df = _handle_suhu_outliers(df, numeric_cols)

    # 5C. Hapus kolom all-null
    all_null_cols = [c for c in df.columns if df[c].isnull().all()]
    if all_null_cols:
        df = df.drop(columns=all_null_cols)
    numeric_cols = [c for c in numeric_cols if c in df.columns]

    # 5D. Imputasi numerik
    df = _impute_numeric(df, numeric_cols)

    # 5D. Imputasi kategorik
    df = _impute_categorical(df)

    # 6. Feature engineering
    df = _feature_engineering(df)

    # 6B. Fallback rule-based labeling jika tidak ada warna kuning di Excel
    # (operator belum menandai baris defect secara manual)
    if len(yellow_rows) == 0:
        df = _rulebased_defect_label(df)

    # 7. Encode kategorik
    df = _encode_categorical(df)

    # 8. Final missing cleanup
    df = _final_missing_cleanup(df)

    # Hapus kolom helper excel_row sebelum export
    df_output = df.drop(columns=["excel_row"], errors="ignore")

    # Tambah row_no untuk traceability (digunakan modelling.py)
    if "row_no" not in df_output.columns:
        df_output["row_no"] = np.arange(1, len(df_output) + 1)

    # Simpan jika diminta
    if save_path:
        df_output.to_csv(save_path, index=False, encoding="utf-8-sig")

    return df_output


# QUICK TEST (jalankan langsung: python preprocessing.py)
if __name__ == "__main__":
    import sys, os

    test_file = (
        sys.argv[1]
        if len(sys.argv) > 1
        else "Rekap_Posisi_Solid_Granulasi_lt_2_26_MARET_2026.xlsx"
    )

    if not os.path.exists(test_file):
        print(f"File tidak ditemukan: {test_file}")
        sys.exit(1)

    print(f"Menjalankan preprocessing pada: {test_file}")
    df_clean = run_preprocessing(test_file, save_path="outputs/dataset_clean.csv")

    print("\n" + "=" * 50)
    print("RINGKASAN DATASET BERSIH")
    print("=" * 50)
    print(f"Total baris    : {len(df_clean)}")
    print(f"Total kolom    : {len(df_clean.columns)}")
    print(f"Baris NORMAL   : {(df_clean['is_defect'] == 0).sum()}")
    print(f"Baris DEFECT   : {(df_clean['is_defect'] == 1).sum()}")
    print(f"Missing value  : {df_clean.isnull().sum().sum()}")
    print(f"Output         : outputs/dataset_clean.csv")